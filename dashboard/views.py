from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta, datetime, date as date_type
from forms_builder.models import FormSubmission, FormTemplate, ItemResponse
from accounts.models import User, Outlet, Department


def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value.strip(), '%Y-%m-%d').date()
    except (ValueError, TypeError, AttributeError):
        return None


def _apply_filters(submissions, request, params, outlet=None, department=None):
    """Apply report filters. outlet/department scope pre-applied by caller."""
    import pytz
    from django.utils import timezone as tz

    staff_id        = params.get('staff')
    form_id         = params.get('form')
    dept_id         = params.get('dept')
    date_from       = _parse_date(params.get('date_from'))
    date_to         = _parse_date(params.get('date_to'))
    response_filter = params.get('response')

    # Scope by outlet if provided
    if outlet:
        submissions = submissions.filter(outlet=outlet)

    # Scope by department if provided
    if department:
        dept_ids = [department.pk] + list(department.children.values_list('pk', flat=True))
        submissions = submissions.filter(department_id__in=dept_ids)

    # Non-admin users: scope to own outlet + dept
    if not request.user.is_admin:
        if not request.user.can_view_all_reports:
            submissions = submissions.filter(submitted_by=request.user)
        else:
            visible = request.user.get_visible_departments()
            if visible is not None:
                submissions = submissions.filter(department_id__in=visible)
            if request.user.outlet:
                submissions = submissions.filter(outlet=request.user.outlet)

    if staff_id:
        try: submissions = submissions.filter(submitted_by_id=int(staff_id))
        except (ValueError, TypeError): pass

    if form_id:
        try: submissions = submissions.filter(form_id=int(form_id))
        except (ValueError, TypeError): pass

    if dept_id:
        try: submissions = submissions.filter(submitted_by__department_id=int(dept_id))
        except (ValueError, TypeError): pass

    if date_from or date_to:
        from datetime import time as time_type
        try:
            local_tz = pytz.timezone('Africa/Harare')
        except Exception:
            local_tz = tz.get_current_timezone()
        if date_from:
            submissions = submissions.filter(submitted_at__gte=local_tz.localize(datetime.combine(date_from, time_type.min)))
        if date_to:
            submissions = submissions.filter(submitted_at__lte=local_tz.localize(datetime.combine(date_to, time_type.max)))

    outlet_id = params.get('outlet')
    if outlet_id and request.user.is_admin:
        try: submissions = submissions.filter(outlet_id=int(outlet_id))
        except (ValueError, TypeError): pass
    dept_filter_id = params.get('dept')
    if dept_filter_id:
        try: submissions = submissions.filter(department_id=int(dept_filter_id))
        except (ValueError, TypeError): pass
    if response_filter == 'no':
        submissions = submissions.filter(no_responses__gt=0)
    elif response_filter == 'perfect':
        submissions = submissions.filter(no_responses=0)

    return submissions


# ── Home — branches for admin, personal for staff ────────────────

@login_required
def home(request):
    if request.user.is_admin:
        return outlet_overview(request)
    else:
        return staff_home(request)


def outlet_overview(request):
    """Admin landing — shows 3 outlet cards with live stats."""
    outlets = Outlet.objects.filter(is_active=True).order_by('order')
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())

    outlet_data = []
    for outlet in outlets:
        subs = FormSubmission.objects.filter(status='submitted', outlet=outlet)
        total      = subs.count()
        today_cnt  = subs.filter(submitted_at__date=today).count()
        flagged    = subs.filter(no_responses__gt=0, submitted_at__date__gte=week_start).count()
        staff_cnt  = User.objects.filter(outlet=outlet, is_active=True).count()
        outlet_data.append({
            'outlet':     outlet,
            'total':      total,
            'today':      today_cnt,
            'flagged':    flagged,
            'staff_count':staff_cnt,
        })

    # Overall totals
    overall_total   = FormSubmission.objects.filter(status='submitted').count()
    overall_today   = FormSubmission.objects.filter(status='submitted', submitted_at__date=today).count()
    overall_flagged = FormSubmission.objects.filter(status='submitted', no_responses__gt=0, submitted_at__date__gte=week_start).count()
    total_staff     = User.objects.filter(is_active=True).count()

    return render(request, 'dashboard/outlet_overview.html', {
        'outlet_data':      outlet_data,
        'overall_total':    overall_total,
        'overall_today':    overall_today,
        'overall_flagged':  overall_flagged,
        'total_staff':      total_staff,
    })


@login_required
def outlet_detail(request, outlet_pk):
    """Drill into one outlet — show departments with stats."""
    if not request.user.is_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')

    outlet = get_object_or_404(Outlet, pk=outlet_pk, is_active=True)
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())

    # Top-level departments for this outlet
    top_depts = Department.objects.filter(outlet=outlet, parent__isnull=True, is_active=True).order_by('order', 'name')

    dept_data = []
    for dept in top_depts:
        # Include dept and all its children
        dept_ids = [dept.pk] + list(dept.children.values_list('pk', flat=True))
        subs = FormSubmission.objects.filter(
            status='submitted', department_id__in=dept_ids, outlet=outlet
        )
        total    = subs.count()
        today_c  = subs.filter(submitted_at__date=today).count()
        flagged  = subs.filter(no_responses__gt=0, submitted_at__date__gte=week_start).count()
        staff_c  = User.objects.filter(department_id__in=dept_ids, outlet=outlet, is_active=True).count()
        children = dept.children.filter(is_active=True).order_by('order', 'name')
        dept_data.append({
            'dept':    dept,
            'total':   total,
            'today':   today_c,
            'flagged': flagged,
            'staff':   staff_c,
            'children':children,
        })

    outlet_total   = FormSubmission.objects.filter(status='submitted', outlet=outlet).count()
    outlet_today   = FormSubmission.objects.filter(status='submitted', submitted_by__outlet=outlet, submitted_at__date=today).count()
    outlet_flagged = FormSubmission.objects.filter(status='submitted', submitted_by__outlet=outlet, no_responses__gt=0).count()
    outlet_staff   = User.objects.filter(outlet=outlet, is_active=True).count()

    return render(request, 'dashboard/outlet_detail.html', {
        'outlet':        outlet,
        'dept_data':     dept_data,
        'outlet_total':  outlet_total,
        'outlet_today':  outlet_today,
        'outlet_flagged':outlet_flagged,
        'outlet_staff':  outlet_staff,
    })


@login_required
def outlet_dept_reports(request, outlet_pk, dept_pk):
    """Reports for a specific department within an outlet."""
    if not request.user.is_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')

    outlet = get_object_or_404(Outlet, pk=outlet_pk)
    dept   = get_object_or_404(Department, pk=dept_pk, outlet=outlet)

    params = request.GET
    base_qs = FormSubmission.objects.filter(status='submitted').select_related('form', 'submitted_by')
    submissions = _apply_filters(base_qs, request, params, outlet=outlet, department=dept)
    submissions = submissions.order_by('-submitted_at')

    total   = submissions.count()
    flagged = submissions.filter(no_responses__gt=0).count()
    perfect = submissions.filter(no_responses=0).count()

    flagged_items = ItemResponse.objects.filter(
        value__iexact='no', submission__in=submissions
    ).select_related('submission__submitted_by', 'submission__form', 'item__section')

    staff_list = User.objects.filter(outlet=outlet, department__in=[dept.pk]+list(dept.children.values_list('pk', flat=True)), is_active=True).order_by('first_name')
    forms = FormTemplate.objects.filter(is_active=True).order_by('name')

    return render(request, 'dashboard/outlet_dept_reports.html', {
        'outlet': outlet, 'dept': dept,
        'submissions': submissions[:100],
        'total': total, 'flagged': flagged, 'perfect': perfect,
        'flagged_items': flagged_items[:200],
        'staff_list': staff_list, 'forms': forms,
        'f_staff':     params.get('staff', ''),
        'f_form':      params.get('form', ''),
        'f_date_from': params.get('date_from', ''),
        'f_date_to':   params.get('date_to', ''),
        'f_response':  params.get('response', ''),
    })


def staff_home(request):
    """Personal dashboard for non-admin users."""
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    my_subs = FormSubmission.objects.filter(submitted_by=request.user)
    total   = my_subs.filter(status='submitted').count()
    today_c = my_subs.filter(status='submitted', submitted_at__date=today).count()
    drafts  = my_subs.filter(status='draft').count()
    flagged = ItemResponse.objects.filter(
        submission__submitted_by=request.user, submission__status='submitted',
        value__iexact='no', submission__submitted_at__date__gte=week_start
    ).count()
    recent  = my_subs.select_related('form').order_by('-created_at')[:8]
    return render(request, 'dashboard/home.html', {
        'is_full_dashboard': False,
        'total_submissions': total,
        'today_submissions': today_c,
        'drafts': drafts,
        'flagged_count': flagged,
        'recent_submissions': recent,
    })


# ── Global reports (non-outlet-scoped, for non-admin users) ──────

@login_required
def reports(request):
    if not request.user.can_view_reports:
        messages.error(request, 'You do not have permission to view reports.')
        return redirect('dashboard:home')

    params = request.GET
    base_qs = FormSubmission.objects.filter(status='submitted').select_related('form', 'submitted_by')
    submissions = _apply_filters(base_qs, request, params).order_by('-submitted_at')

    total   = submissions.count()
    flagged = submissions.filter(no_responses__gt=0).count()
    perfect = submissions.filter(no_responses=0).count()

    flagged_items = ItemResponse.objects.filter(
        value__iexact='no', submission__in=submissions
    ).select_related('submission__submitted_by', 'submission__form', 'item__section')

    # Staff list scoped to user's outlet
    if request.user.is_admin:
        staff_list  = User.objects.filter(is_active=True).select_related('outlet', 'department').order_by('first_name', 'username')
        departments = Department.objects.select_related('outlet', 'parent').order_by('outlet__order', 'order', 'name')
        outlets     = Outlet.objects.filter(is_active=True)
    else:
        user_outlet = request.user.outlet
        staff_list  = User.objects.filter(is_active=True, outlet=user_outlet).select_related('department').order_by('first_name') if user_outlet else None
        # Non-admin sees departments within their outlet only
        departments = Department.objects.filter(outlet=user_outlet).select_related('parent').order_by('order', 'name') if user_outlet else None
        outlets     = None

    forms = FormTemplate.objects.filter(is_active=True).order_by('name')

    return render(request, 'dashboard/reports.html', {
        'submissions': submissions[:100],
        'total': total, 'flagged': flagged, 'perfect': perfect,
        'flagged_items': flagged_items[:200],
        'staff_list': staff_list, 'forms': forms,
        'departments': departments, 'outlets': outlets,
        'f_staff':     params.get('staff', ''),
        'f_form':      params.get('form', ''),
        'f_dept':      params.get('dept', ''),
        'f_outlet':    params.get('outlet', ''),
        'f_dept':      params.get('dept', ''),
        'f_date_from': params.get('date_from', ''),
        'f_date_to':   params.get('date_to', ''),
        'f_response':  params.get('response', ''),
            })


@login_required
def delete_submission(request, pk):
    if not request.user.can_delete_submissions:
        messages.error(request, 'You do not have permission to delete submissions.')
        return redirect('dashboard:reports')
    submission = get_object_or_404(FormSubmission, pk=pk)
    if request.method == 'POST':
        form_name = submission.form.name
        submission.delete()
        messages.success(request, f'Submission of "{form_name}" deleted.')
        return redirect('dashboard:reports')
    return render(request, 'dashboard/confirm_delete.html', {'submission': submission})


# ── Export & Print (carry outlet/dept context via GET params) ────

@login_required
def export_excel(request):
    if not request.user.can_view_reports:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'openpyxl is required. Run: pip install openpyxl')
        return redirect('dashboard:reports')
    from django.http import HttpResponse

    submissions = FormSubmission.objects.filter(status='submitted').select_related('form', 'submitted_by__outlet', 'submitted_by__department')
    submissions = _apply_filters(submissions, request, request.GET).order_by('-submitted_at')

    flagged_items = ItemResponse.objects.filter(
        value__iexact='no', submission__in=submissions
    ).select_related('submission__submitted_by', 'submission__form', 'item__section').order_by('-submission__submitted_at')

    wb = openpyxl.Workbook()
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', fgColor='1A1200')
    alt_fill     = PatternFill('solid', fgColor='F0F4F8')
    flag_fill    = PatternFill('solid', fgColor='FFF0F0')
    center       = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left         = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border  = Border(
        left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'),
    )

    def style_header_row(ws, cols):
        for col_idx, (header, width) in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = center; cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 22

    def style_data_cell(cell, row_idx, is_flag=False):
        cell.alignment = left; cell.border = thin_border
        if is_flag: cell.fill = flag_fill
        elif row_idx % 2 == 0: cell.fill = alt_fill

    ws1 = wb.active
    ws1.title = 'Submissions'
    ws1.freeze_panes = 'A2'
    cols1 = [
        ('Outlet',14),('Form',28),('Staff Member',20),('Department',18),
        ('Branch',14),('Shift',12),('Date',18),('Time',10),
        ('Completion %',14),('Issues',12),('Status',12),
    ]
    style_header_row(ws1, cols1)
    for row_idx, sub in enumerate(submissions, 2):
        is_flag = sub.no_responses > 0
        data = [
            sub.submitted_by.outlet_name,
            sub.form.name, sub.submitted_by.display_name, sub.submitted_by.dept_name,
            sub.submitted_by.outlet_name, sub.get_shift_display(),
            sub.submitted_at.strftime('%d %b %Y') if sub.submitted_at else '',
            sub.submitted_at.strftime('%H:%M') if sub.submitted_at else '',
            sub.completion_percentage, sub.no_responses,
            'Issues Found' if is_flag else 'Clean',
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            style_data_cell(cell, row_idx, is_flag)
    ws1.auto_filter.ref = f'A1:{get_column_letter(len(cols1))}1'

    ws2 = wb.create_sheet('Flagged Items')
    ws2.freeze_panes = 'A2'
    cols2 = [('Checklist Item',36),('Section',22),('Form',24),('Staff Member',20),('Outlet',16),('Date',16),('Comment',40),('Photo',14)]
    style_header_row(ws2, cols2)
    for row_idx, r in enumerate(flagged_items, 2):
        data = [
            r.item.label, r.item.section.title, r.submission.form.name,
            r.submission.submitted_by.display_name, r.submission.submitted_by.outlet_name,
            r.submission.submitted_at.strftime('%d %b %Y') if r.submission.submitted_at else '',
            r.comment or '', 'Yes' if r.image else 'No',
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = flag_fill; cell.alignment = left; cell.border = thin_border
    ws2.auto_filter.ref = f'A1:{get_column_letter(len(cols2))}1'

    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer); buffer.seek(0)
    filename = f"Ambassador_Report_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def print_report(request):
    if not request.user.can_view_reports:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    submissions = FormSubmission.objects.filter(status='submitted').select_related('form', 'submitted_by')
    submissions = _apply_filters(submissions, request, request.GET).order_by('-submitted_at')
    flagged_items = ItemResponse.objects.filter(value__iexact='no', submission__in=submissions).select_related('submission__submitted_by', 'submission__form', 'item__section')
    total   = submissions.count()
    flagged = submissions.filter(no_responses__gt=0).count()
    perfect = submissions.filter(no_responses=0).count()
    return render(request, 'dashboard/print_report.html', {
        'submissions': submissions, 'flagged_items': flagged_items,
        'total': total, 'flagged': flagged, 'perfect': perfect,
        'date_from': request.GET.get('date_from', ''), 'date_to': request.GET.get('date_to', ''),
        'generated_by': request.user.display_name, 'generated_at': timezone.now(),
    })
